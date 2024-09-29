import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from utils import *
import warnings

# Supress all warnings
warnings.filterwarnings('ignore')

def avails_process():
    print('Processing avails...')

    app_dir = get_app_dir()

    latam_countries = [
        'Costa Rica',
        'El Salvador',
        'Guatemala',
        'Honduras',
        'Nicaragua',
        'Panama',
        'Mexico',
        'Argentina',
        'Bolivia',
        'Brazil',
        'Chile',
        'Colombia',
        'Ecuador',
        'Paraguay',
        'Peru',
        'Uruguay',
        'Venezuela'
    ]

    # Load the data
    windows_df = pd.read_pickle(os.path.join(app_dir, 'data', 'tables', 'windows.pkl'))
    contracts = pd.read_pickle(os.path.join(app_dir, 'data', 'tables',  'contracts.pkl'))
    titles = pd.read_pickle(os.path.join(app_dir, 'data', 'tables', 'titles.pkl'))
    roles = pd.read_pickle(os.path.join(app_dir, 'data', 'tables', 'roles.pkl'))
    people = pd.read_pickle(os.path.join(app_dir, 'data', 'tables', 'people.pkl'))
    
    # Filter contracts_df for 'Normal' status
    contracts_filtered = contracts[contracts['status'] == 'Normal']

    # Perform an inner join on the 'contract' column
    windows_df = pd.merge(windows_df, contracts_filtered[['contract', 'contract_type', 'distributor', 'status']], on='contract', how='inner')

    # Use the window column as the index
    windows_df.set_index('window', inplace=True)

    # Filter titles for 'SD Tape' and 'SD File' in the 'original_format' column
    titles = titles.loc[
        ~(titles['original_format'].isin(['SD Tape', 'SD File']))
        ]

    titles.drop([
        'imdb_code', 
        'project_group', 
        'adj_running_time', 
        'aka_2', 
        'title_code'
        ], axis=1, inplace=True)

    # Create a talent dataframe by merging the roles and people dataframes
    talent = pd.merge(
        roles[['title', 'role', 'person']], 
        people[['person', 'name']],
        on='person', 
        how='inner'
    )

    # Group the talent dataframe by 'title' and 'role' and aggregate the 'name' column as a list
    talent_list = talent.groupby(['title', 'role'])[['name']].agg({'name':list})
    talent_list = talent_list.unstack(1)
    talent_list.columns = [role.lower() for _, role in talent_list.columns]

    # Merge the titles dataframe with the talent_list dataframe
    titles = titles.join(
        talent_list, 
        how='left',
    )

    # Create a sales_activity dataframe by filtering the windows_df dataframe for 'Sales' and 'Holdback' in the 'contract_type' and 'license_type' columns respectively
    sales_activity = windows_df.loc(axis=0)[
        (windows_df['contract_type'] == 'Sales') &
        (windows_df['license_type'] != 'Holdback')
    ]

    # Group the sales_activity dataframe by 'title', 'group', 'right_name', 'country_name', 'license_type', and 'distributor' and aggregate the 'end_date' column as the maximum value
    sales_activity = sales_activity.groupby([
        'title',
        'group',
        'right_name',
        'country_name',
        'license_type', 
        'distributor'
    ], dropna=False)['end_date'].max()

    # Unstack the sales_activity dataframe by the last two levels
    sales_activity = sales_activity.unstack(level=[-1, -2])

    # Drop rows with all NaN values
    sales_activity = sales_activity[~sales_activity.isna().all(axis=1)]
    sales_cols = sales_activity.columns
    
    # sort sales_cols alphabetically
    sales_cols = sorted(sales_cols, key=lambda x: x[0])
    sales_activity = sales_activity[sales_cols]

    # Create windows_mat by grouping the windows_df dataframe by 'title', 'group', 
    # 'right_name', 'country_name', 'contract_type', and 'license_type' and 
    # aggregating the 'start_date' and 'end_date' columns as the maximum value
    window_mat = windows_df.groupby([
        'title',
        'group',
        'right_name',
        'country_name',
        'contract_type',
        'license_type'
    ], dropna=False).agg({
        'start_date': np.max,
        'end_date': np.max
    })

    # Add 1 day to the end_date to make it inclusive
    window_mat.loc[
        pd.IndexSlice[:,:,:,:,'Sales', :], 
        'end_date'
        ] = window_mat.loc[
            pd.IndexSlice[:,:,:,:,'Sales', :], 
            'end_date'
            ] + pd.Timedelta(1, 'D')

    # Unstack the window_mat dataframe by the last two levels
    window_mat = window_mat.unstack(level=[-2, -1])
    window_mat['today'] = pd.Timestamp('today').date()

    # Calculate the exclusive avail start date
    st_date_excl = window_mat.loc[
        (~window_mat[('start_date',  'Acquisition',       'License')].isna()) &
        (window_mat[(  'end_date',  'Acquisition',       'License')] >= pd.Timestamp('today') + pd.Timedelta(26, 'W')) 
    ][[
        ('start_date',  'Acquisition',       'License'),
        (  'end_date',  'Acquisition',      'Holdback'),
        (  'end_date', 'Sales', 'Non-Exclusive'),
        (  'end_date', 'Sales',       'License'),
        (  'end_date', 'Sales',      'Holdback'),
        (     'today',    '',              ''),
    ]]
    st_date_excl = st_date_excl.fillna(pd.Timestamp.min)
    st_date_excl_idx = st_date_excl.index

    # Create an avails_df dataframe with the exclusive column
    avails_df = pd.DataFrame(
        st_date_excl.values.max(axis=1), 
        columns=['exclusive'], 
        index=st_date_excl_idx
    )

    # Calculate the non-exclusive start date
    st_date_non_excl = window_mat.loc[
        (~window_mat[('start_date',  'Acquisition',       'License')].isna() |
        (~window_mat[('start_date',  'Acquisition',       'Non-Exclusive')].isna())) &
        ((window_mat[(  'end_date',  'Acquisition',       'License')] >= pd.Timestamp('today') + pd.Timedelta(26, 'W'))  |
        (window_mat[(  'end_date',  'Acquisition',       'Non-Exclusive')] >= pd.Timestamp('today') + pd.Timedelta(26, 'W')))
    ][[
        ('start_date',  'Acquisition',       'License'),
        ('start_date',  'Acquisition', 'Non-Exclusive'),
        (  'end_date',  'Acquisition',      'Holdback'),
        (  'start_date', 'Sales',       'License'),
        (  'end_date', 'Sales',       'License'),
        (  'end_date', 'Sales',      'Holdback'),
        (     'today',    '',              ''),
    ]]
    st_date_non_excl = st_date_non_excl.fillna(pd.Timestamp.min)
    st_date_non_excl['non-exclusive'] = st_date_non_excl.apply(max_date, axis=1)

    # Join the non-exclusive column of st_date_non_excl with the avails_df dataframe
    avails_df = avails_df.join(st_date_non_excl[['non-exclusive']], how='outer')

    avails_df.columns = [''.join(col) for col in avails_df.columns]

    # Calculate the acquitsition expiry date
    acq_exp = window_mat.loc[
        (~window_mat[('start_date',  'Acquisition',       'License')].isna() |
        (~window_mat[('start_date',  'Acquisition',       'Non-Exclusive')].isna())) &
        ((window_mat[(  'end_date',  'Acquisition',       'License')] >= pd.Timestamp('today') + pd.Timedelta(26, 'W'))  |
        (window_mat[(  'end_date',  'Acquisition',       'Non-Exclusive')] >= pd.Timestamp('today') + pd.Timedelta(26, 'W')))
    ][[
        (  'end_date',  'Acquisition',       'License'),
        (  'end_date',  'Acquisition', 'Non-Exclusive'),
    ]]
    acq_exp = acq_exp.fillna(pd.Timestamp.min)
    avails_df['acq_expires'] = acq_exp.values.max(axis=1)

    # Calculate the non-exclusive end date
    non_excl_end_date = avails_df[['non-exclusive', 'acq_expires']].join(
        window_mat[[
            ('start_date',       'Sales',       'License'),
            ('start_date',       'Sales',      'Holdback'),
        ]], 
        how='left'
    )
    avails_df['non-exclusive_end_date'] = non_excl_end_date.apply(non_exclusive_end_date, axis=1)
    avails_df = avails_df.reset_index().groupby(['title', 'group', 'country_name'])[['exclusive', 'non-exclusive', 'acq_expires', 'non-exclusive_end_date']].max()

    # Replace the exclusive and non-exclusive columns of avails_df where the boolean condition 
    # avails_df['acq_expires'] - avails_df['exclusive'] < pd.Timedelta(26, 'W') is True with NaT
    avails_df.loc[avails_df['acq_expires'] - avails_df['exclusive'] < pd.Timedelta(26, 'W'), 'exclusive'] = pd.NaT
    avails_df.loc[avails_df['acq_expires'] - avails_df['non-exclusive'] < pd.Timedelta(26, 'W'), 'non-exclusive'] = pd.NaT

    # Drop the last column of st_date_non_excl
    non_exclusive_dates = st_date_non_excl.drop(st_date_non_excl.columns[-1], axis=1).max(axis=1)

    # name the non_exclusive_dates series as 'non-exclusive'
    non_exclusive_dates.name = 'non-exclusive'

    # Create a non_exclusive_dates dataframe with the 'non-exclusive' column and a 'non-exclusive_end_date' column with NaT values
    non_exclusive_dates = pd.DataFrame(non_exclusive_dates)
    non_exclusive_dates['non-exclusive_end_date'] = pd.NaT

    # Group the non_exclusive_dates dataframe by 'title', 'group', and 'country_name' and aggregate the maximum value
    non_exclusive_dates = non_exclusive_dates.groupby(['title', 'group', 'country_name']).max()

    # Replace the non-exclusive and non-exclusive_end_date columns of avails_df where the boolean condition 
    # avails_df['non-exclusive_end_date'] - avails_df['non-exclusive'] < pd.Timedelta(26, 'W') is True 
    # with the corresponding values of the non_exclusive_dates series where the index matches
    avails_df.loc[avails_df['non-exclusive_end_date'] - avails_df['non-exclusive'] < pd.Timedelta(26, 'W'), ['non-exclusive', 'non-exclusive_end_date']] = non_exclusive_dates

    # Drop rows with all NaT values
    avails_df.dropna(subset=['exclusive', 'non-exclusive'], how='all', inplace=True)

    # Create a slice of avails_df with the 'Premium Pay TV (Local)' and 'Premium Pay TV (Pan Regional)' rights
    latam_ptv_slice = avails_df.loc[
        pd.IndexSlice[:, ['Premium Pay TV (Local)', 'Premium Pay TV (Pan Regional)'], latam_countries],
        ['exclusive', 'non-exclusive']
    ]

    # Group the latam_ptv_slice dataframe by 'title' and aggregate the maximum value
    max_latam_ptv_dates = latam_ptv_slice.reset_index().groupby(['title'])[['exclusive', 'non-exclusive']].max()
    latam_pan_ptv_slice = avails_df.loc[
        pd.IndexSlice[:, 'Premium Pay TV (Pan Regional)', latam_countries],
        ['exclusive', 'non-exclusive']
    ]

    # Merge the exclusive and non-exclusive columns of a slice of avails_df with max_latam_ptv_dates by title
    latam_pan_ptv_slice = latam_pan_ptv_slice.merge(
        max_latam_ptv_dates,
        how='inner',
        left_index=True,
        right_index=True,
        suffixes=('_prev', '')
    )
    latam_pan_ptv_slice.drop(['exclusive_prev', 'non-exclusive_prev'], axis=1, inplace=True)

    # Replace the exclusive and non-exclusive columns of avails_df with the latam_pan_ptv_slice
    avails_df.loc[
        pd.IndexSlice[:, 'Premium Pay TV (Pan Regional)', latam_countries],
        ['exclusive', 'non-exclusive']
    ] = latam_pan_ptv_slice

    # Create a slice of avails_df with the 'Basic Pay TV (Local)' and 'Basic Pay TV (Pan Regional)' rights
    latam_bptv_slice = avails_df.loc[
        pd.IndexSlice[:, ['Basic Pay TV (Local)', 'Basic Pay TV (Pan Regional)'], latam_countries],
        ['exclusive', 'non-exclusive']
    ]

    # Group the latam_bptv_slice dataframe by 'title' and aggregate the maximum value
    max_latam_bptv_dates = latam_bptv_slice.reset_index().groupby(['title'])[['exclusive', 'non-exclusive']].max()
    
    # Create a slice of avails_df with the 'Basic Pay TV (Pan Regional)' rights
    latam_pan_bptv_slice = avails_df.loc[
        pd.IndexSlice[:, 'Basic Pay TV (Pan Regional)', latam_countries],
        ['exclusive', 'non-exclusive']
    ]

    # Merge the exclusive and non-exclusive columns of a slice of avails_df with max_latam_bptv_dates by title
    latam_pan_bptv_slice = latam_pan_bptv_slice.merge(
        max_latam_bptv_dates,
        how='left',
        left_index=True,
        right_index=True,
        suffixes=('_prev', '')
    )
    latam_pan_bptv_slice.drop(['exclusive_prev', 'non-exclusive_prev'], axis=1, inplace=True)

    # Replace the exclusive and non-exclusive columns of avails_df with the latam_pan_bptv_slice
    avails_df.loc[
        pd.IndexSlice[:, 'Basic Pay TV (Pan Regional)', latam_countries],
        ['exclusive', 'non-exclusive']
    ] = latam_pan_bptv_slice

    useful_rights = [
        'SVOD', 
        'Premium Pay TV (Pan Regional)',
        'Premium Pay TV (Local)',
        'Basic Pay TV (Pan Regional)', 
        'Basic Pay TV (Local)', 
        'AVOD', 
    ]

    # Create a slice of avails_df with the useful_rights
    avails_redux = avails_df.loc[pd.IndexSlice[:,useful_rights,:,:,:]]
    
    # Join the avails_redux dataframe with the sales_activity dataframe
    avails_redux = avails_redux.join(sales_activity, how='left')
    avails_redux.reset_index(inplace=True)

    # Create a 'first_run_status' column in avails_redux with the value 'First Run' if all the values in the sales_cols are NaN, otherwise 'Library'
    avails_redux['first_run_status'] = avails_redux[sales_cols].isna().all(axis=1).map({True: 'First Run', False: 'Library'})

    # Group the avails_redux dataframe by 'title', 'group', and 'country_name' and aggregate the 'first_run_status' column as a set
    status_df = avails_redux.groupby(['title', 'group', 'country_name'])['first_run_status'].apply(set)

    # Replace sets with length greater than 1 with a singleton set containing 'Library'
    status_df = status_df.apply(lambda x: {'Library'} if len(x) > 1 else x)

    # Convert sets to strings
    status_df = status_df.apply(lambda x: ''.join(x))

    # Convert the 'first_run_status' column of status_df to a dataframe
    status_df = status_df.to_frame()

    # Unstack the status_df dataframe by the 'group' column
    status_df.columns = ['first_run_status']
    status_df = status_df.unstack('group')

    # delete the multiindex
    status_df.columns = status_df.columns.droplevel()
    
    # if Basic Pay TV (Local) is Library, then set Basic Pay TV (Pan Regional) to Library and vice versa
    status_df.loc[status_df['Basic Pay TV (Pan Regional)'] == 'Library', 'Basic Pay TV (Local)'] = 'Library'
    status_df.loc[status_df['Basic Pay TV (Local)'] == 'Library', 'Basic Pay TV (Pan Regional)'] = 'Library'

    # if Premium Pay TV (Local) is Library, then set Premium Pay TV (Pan Regional) to Library and vice versa
    status_df.loc[status_df['Premium Pay TV (Pan Regional)'] == 'Library', 'Premium Pay TV (Local)'] = 'Library'
    status_df.loc[status_df['Premium Pay TV (Local)'] == 'Library', 'Premium Pay TV (Pan Regional)'] = 'Library'

    # if Basic Pay TV (Local) is Library, then set Premium Pay TV (Local) and Premium Pay TV (Pan Regional) to Library
    status_df.loc[status_df['Basic Pay TV (Local)'] == 'Library', 'Premium Pay TV (Local)'] = 'Library'
    status_df.loc[status_df['Basic Pay TV (Local)'] == 'Library', 'Premium Pay TV (Pan Regional)'] = 'Library'

    # if SVOD is Library, then set Premium Pay TV (Local) and Premium Pay TV (Pan Regional) to Library
    status_df.loc[status_df['SVOD'] == 'Library', 'Premium Pay TV (Local)'] = 'Library'
    status_df.loc[status_df['SVOD'] == 'Library', 'Premium Pay TV (Pan Regional)'] = 'Library'

    # stack group back to columns
    status_df = status_df.stack('group')

    # name the status_df series 'first_run_status'
    status_df.name = 'first_run_status'

    status_df = status_df.reset_index()

    # Merge the status_df dataframe with the avails_redux dataframe
    avails_redux = avails_redux.merge(status_df, how='left', on=['title', 'group', 'country_name'], suffixes=('_2drop', ''))
    avails_redux = avails_redux.drop(columns=['first_run_status_2drop'])

    # Group the avails_redux dataframe by 'title', 'group', 'exclusive', and 'non-exclusive' and aggregate the 'right_name', 
    # 'country_name', and 'first_run_status' columns as a set
    groupby_cols = [
            'title', 
            'group', 
            'exclusive', 
            'non-exclusive',
    ]
    agg_cols = [
        'right_name', 
        'country_name',
        'first_run_status'
    ]
    agg_dict = {ag_col:set for ag_col in agg_cols}
    other_cols = [col for col in avails_redux.columns if (not col in groupby_cols) and (not col in agg_cols)]
    apply_funcs = {**agg_dict, **{col:'first' for col in other_cols}}

    # Aggregate the avails_redux dataframe: output_audit
    output_audit = avails_redux.groupby(groupby_cols, dropna=False)[agg_cols+other_cols].agg(apply_funcs)
    output_audit['first_run_status'] = output_audit['first_run_status'].apply(lambda x: {'Library'} if len(x) > 1 else x)
    output_audit['first_run_status'] = output_audit['first_run_status'].apply(lambda x: ''.join(x))
    output_audit.reset_index(inplace=True)

    # Check if countries_to_check is a subset of the countries in each row
    output_audit['avails_region'] = output_audit['country_name'].apply(avails_region)
    output_audit = output_audit.loc[output_audit['avails_region'].apply(lambda x: len(x) > 0)]

    for col in output_audit.select_dtypes(include=['object']).columns:
        output_audit[col] = output_audit[col].apply(clean_str)

    other_cols = [col for col in output_audit.columns if col not in sales_cols]

    # Merge the output_audit dataframe with the titles dataframe
    output_audit = output_audit.merge(
        titles.reset_index(),
        how='inner',
        on='title'
    )
    output_audit.drop('title', axis=1, inplace=True)
    output_audit.rename(columns={'name':'title'}, inplace=True)
    output_audit.columns = [
        '_'.join(col) if col[0] != '' and type(col) != str else col for col in output_audit.columns
    ]

    # format 'year_completed' column as an integer
    output_audit['year_completed'] = output_audit['year_completed'].apply(lambda x: int(x))

    # sort values of output_audit by year_completed in descending order
    output_audit.sort_values('year_completed', ascending=False, inplace=True)

    # define the columns to be formatted as dates
    date_cols = [
        'non-exclusive',
        'exclusive',
        'non-exclusive_end_date',
        'acq_expires'
    ]

    sales_cols = [
        '_'.join(col) for col in sales_cols
    ]

    cols_order = [
            'Title',   
            'Region',
            'Year',
            'Genre',
            'Rights Group',
            'First Run / Library',
            'Non-Exclusive Start Date',
            'Non-Exclusive End Date',
            'Exclusive Start Date',
            'Acq_expires',
            'Original Language',
            'Dialogue Language',
            'Subtitle Language',
            'USA Rating',
            'Running Time',
            'cast',
            'director',
            'synopsis',
            'Website (Trailer)',
            'Link (full movie)',
            'Password',
            'IMDB',
    ] + sales_cols + [
            'us_box_office',
            'latam_box_office',
            'rating_usa',
            'rating_mexico',
            'rating_brazil',
            'rating_argentina',
            'rating_bolivia',
            'rating_chile',
            'rating_colombia',
            'rating_costa_rica',
            'rating_ecuador',
            'rating_el_salvador',
            'rating_guatemala',
            'rating_honduras',
            'rating_nicaragua',
            'rating_panama',
            'rating_paraguay',
            'rating_peru',
            'rating_dominican_republic',
            'rating_uruguay',
            'rating_venezuela'
    ]
        
    # create a dictionary with the column names as keys and the lowercased column names as values, unless the column name is in sales_cols. If the column name is in sales_cols, then the value is the column name itself
    col_dict = {}

    for col in cols_order:
        if col in sales_cols:
            col_dict[col] = col
        else:
            col_dict[col] = col.lower()


    col_dict['Region'] = 'avails_region'
    col_dict['Rights Group'] = 'group'
    col_dict['First Run / Library'] = 'first_run_status'
    col_dict['Non-Exclusive Start Date'] = 'non-exclusive'
    col_dict['Exclusive Start Date'] = 'exclusive'
    col_dict['Non-Exclusive End Date'] = 'non-exclusive_end_date'
    col_dict['Acq_expires'] = 'acq_expires'
    col_dict['Original Language'] = 'original_language'
    col_dict['Dialogue Language'] = 'dialogue_language'
    col_dict['Subtitle Language'] = 'subtitle_language'
    col_dict['USA Rating'] = 'rating_usa'
    col_dict['Running Time'] = 'running_time'
    col_dict['Website (Trailer)'] = 'website'
    col_dict['Link (full movie)'] = 'link'
    col_dict['Year'] = 'year_completed'

    # Swap the keys and values of col_dict
    col_dict = {v: k for k, v in col_dict.items()}

    # format the date columns and sales columns as dates. If the value is NaT, then do not format it
    for col in date_cols + sales_cols:
        output_audit[col] = output_audit[col].apply(
            lambda x: pd.to_datetime(x, errors='coerce').date() if x != 'NaT' else x
        )

    # Rename the columns of output_audit using col_dict
    output_audit.rename(columns=col_dict, inplace=True)

    # get the values of col_dict as a list
    col_list = list(col_dict.values())

    # Create an 'avails' directory if it does not exist
    if not os.path.exists(os.path.join(app_dir, 'avails')):
        os.makedirs(os.path.join(app_dir, 'avails'))

    # Export output_audit[col_list] to excel and format the columns
    output_audit[col_list].dropna(axis=1, how='all').to_excel(os.path.join(app_dir, 'avails', 'avails_audit.xlsx'), index=False)

    # Format the excel sheet
    wb2 = load_workbook(os.path.join(app_dir, 'avails', 'avails_audit.xlsx'))
    ws2 = wb2.active    
    ws2.auto_filter.ref = ws2.dimensions
    ws2.freeze_panes = ws2['C2']
    for col in ws2.iter_cols(min_col=1):
        for cell in col:
            ws2.column_dimensions[cell.column_letter].width = 30
            if cell.column_letter in ['G', 'H', 'I', 'J']:
                cell.number_format = 'YYYY-MM-DD'
    for row in ws2.iter_rows(min_row=1):
        for cell in row:
            cell.font = Font(size=11)
            cell.alignment = Alignment(wrap_text=True)
            ws2.row_dimensions[cell.row].height = 50
    for row in ws2.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            for cell in row:
                if cell.column_letter == 'A':
                    cell.alignment = Alignment(horizontal='left')
                    cell.font = Font(bold=True)

    wb2.save(os.path.join(app_dir, 'avails', 'avails_audit.xlsx'))
    wb2.close()

    # Append the 'country_of_origin' column to col_list
    col_list.append('country_of_origin')

    # Create a combined_df dataframe with the columns in col_list
    combined_df = output_audit[col_list].copy()

    # convet the sales cols of output_audit to a dictionary and drop pd.NaT values
    sales_dict_list = combined_df[sales_cols].to_dict(orient='records')
    sales_dict_list = [
        {k:str(v) for k, v in sales_dict.items() if type(v) != pd._libs.tslibs.nattype.NaTType} for sales_dict in sales_dict_list
    ]
    combined_df['Sales'] = sales_dict_list
    combined_df['Sales'] = combined_df['Sales'].apply(clean_str)

    # Group the combined_df dataframe by 'title' and 'region' and aggregate the minimum value of the 'acq_expires' column
    acq_expires = combined_df[['Title', 'Region', 'Acq_expires']].groupby(['Title', 'Region']).min()

    summary_cols = [
        'Title', 
        'Region', 
        'Rights Group',
        'First Run / Library', 
        'Non-Exclusive Start Date',
        'Non-Exclusive End Date', 
        'Exclusive Start Date',
        'Sales',
    ]

    # Create a combined_unstacked dataframe with the summary_cols
    combined_unstacked = combined_df[summary_cols].set_index(['Title', 'Rights Group', 'Region']).unstack('Rights Group')

    # swap the levels of the column index and sort the columns
    combined_unstacked.columns = combined_unstacked.columns.swaplevel(0,1)
    combined_unstacked.sort_index(axis=1, inplace=True)
    combined_unstacked.columns = [
        '_'.join(col) if col[0] != '' and type(col) != str else col for col in combined_unstacked.columns
    ]

    # Merge the combined_unstacked dataframe with the acq_expires dataframe
    combined_unstacked = combined_unstacked.merge(
        acq_expires,
        how='left',
        left_index=True,
        right_index=True
    )

    other_cols = [col for col in combined_df.columns if col not in summary_cols and col != 'Acq_expires']
    
    # Group the combined_df dataframe by 'title' and 'region' and aggregate the first value of the other_cols: metadata_df
    metadata_df = combined_df.groupby(['Title', 'Region'])[other_cols].first()

    # Merge the combined_unstacked dataframe with the metadata_df dataframe
    combined_unstacked = combined_unstacked.merge(
        metadata_df,
        how='left',
        left_index=True,
        right_index=True
    )

    # clean the strings of the columns that are objects
    for col in combined_unstacked.select_dtypes(include=['object']).columns:
        combined_unstacked[col] = combined_unstacked[col].apply(clean_str)

    combined_unstacked.reset_index(inplace=True)

    # sort combined_unstacked by Year in descending order and then by Title in ascending order
    combined_unstacked.sort_values(['Year', 'Title'], ascending=[False, True], inplace=True)

    # rename the 'cast' colunn in combined_unstacked to 'Cast'
    combined_unstacked.rename(columns={'cast':'Cast'}, inplace=True)
    combined_unstacked.rename(columns={'director':'Director'}, inplace=True)
    combined_unstacked.rename(columns={'rating_usa':'Rating USA'}, inplace=True)
    combined_unstacked.rename(columns={'Acq_expires':'Acq Expires'}, inplace=True)
    combined_unstacked.rename(columns={'country_of_origin':'Country of Origin'}, inplace=True)

    cols_ordered = [
        'Title',
        'Region',
        'Year',
        'Genre',
        'Premium Pay TV (Pan Regional)_First Run / Library',
        'Premium Pay TV (Pan Regional)_Non-Exclusive Start Date',
        'Premium Pay TV (Pan Regional)_Non-Exclusive End Date',
        'Premium Pay TV (Pan Regional)_Exclusive Start Date',
        'Premium Pay TV (Pan Regional)_Sales',
        'Premium Pay TV (Local)_First Run / Library',
        'Premium Pay TV (Local)_Non-Exclusive Start Date',
        'Premium Pay TV (Local)_Non-Exclusive End Date',
        'Premium Pay TV (Local)_Exclusive Start Date',
        'Premium Pay TV (Local)_Sales',
        'Basic Pay TV (Pan Regional)_First Run / Library',
        'Basic Pay TV (Pan Regional)_Non-Exclusive Start Date',
        'Basic Pay TV (Pan Regional)_Non-Exclusive End Date',
        'Basic Pay TV (Pan Regional)_Exclusive Start Date',
        'Basic Pay TV (Pan Regional)_Sales',
        'Basic Pay TV (Local)_First Run / Library',
        'Basic Pay TV (Local)_Non-Exclusive Start Date',
        'Basic Pay TV (Local)_Non-Exclusive End Date',
        'Basic Pay TV (Local)_Exclusive Start Date',
        'Basic Pay TV (Local)_Sales',
        'SVOD_First Run / Library',
        'SVOD_Non-Exclusive Start Date',
        'SVOD_Non-Exclusive End Date',
        'SVOD_Exclusive Start Date',
        'SVOD_Sales',
        'AVOD_First Run / Library',
        'AVOD_Non-Exclusive Start Date',
        'AVOD_Non-Exclusive End Date',
        'AVOD_Exclusive Start Date',
        'AVOD_Sales',
        'Acq Expires',
        'Cast',
        'Director',
        'synopsis',
        'Running Time',
        'Country of Origin',
        'Original Language',
        'Dialogue Language',
        'Subtitle Language',
        'Rating USA',
        'Website (Trailer)',
        'Link (full movie)',
        'Password',
        'IMDB'
    ] + sales_cols + [
        'us_box_office',
        'latam_box_office',
        'rating_mexico',
        'rating_brazil',
        'rating_argentina',
        'rating_bolivia',
        'rating_chile',
        'rating_colombia',
        'rating_costa_rica',
        'rating_ecuador',
        'rating_el_salvador',
        'rating_guatemala',
        'rating_honduras',
        'rating_nicaragua',
        'rating_panama',
        'rating_paraguay',
        'rating_peru',
        'rating_dominican_republic',
        'rating_uruguay',
        'rating_venezuela'
    ]

    # Export the combined_unstacked dataframe to excel and format the columns
    combined_unstacked[cols_ordered].dropna(axis=1, how='all').to_excel(os.path.join(app_dir, 'avails', 'avails.xlsx'), index=False)

    # Format the excel sheet
    wb2 = load_workbook(os.path.join(app_dir, 'avails', 'avails.xlsx'))
    ws2 = wb2.active

    # Apply an autofilter to the first row of 'avails.xlsx'
    ws2.auto_filter.ref = ws2.dimensions

    # Format the height of the rows of 'avails.xlsx' to be higher 
    for row in ws2.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            cell.font = Font(size=11)
            ws2.row_dimensions[cell.row].height = 50

    # Format the width of the columns of 'avails.xlsx' to be wider
    for col in ws2.iter_cols(min_col=1):
        for cell in col:
            ws2.column_dimensions[cell.column_letter].width = 30

    for row in ws2.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            if cell.column_letter == 'A':
                cell.alignment = Alignment(horizontal='left')
                cell.font = Font(bold=True)

    # Freeze the first row of 'avails.xlsx'
    ws2.freeze_panes = ws2['C2']

    # Save the workbook
    wb2.save(os.path.join(app_dir, 'avails', 'avails.xlsx'))
    wb2.close()

    # Free TV avails
    free_tv_avails = avails_df.reset_index()

    # Index slice avails_df where group is 'Free TV' and 'country_name' is in latam_countries
    free_tv_avails = free_tv_avails.loc[
        (free_tv_avails['group'] == 'Free TV') &
        (free_tv_avails['country_name'].isin(latam_countries + ['Dominican Republic', 'Puerto Rico'])),
        ['title', 'country_name', 'group', 'exclusive', 'acq_expires']
    ].drop_duplicates()

    free_tv_avails.drop('group', axis=1, inplace=True)

    # group by 'title' and 'country_name' and return the maximum value of 'exclusive' and 'acq_expires'
    free_tv_avails = free_tv_avails.groupby(['title', 'country_name'])[['exclusive', 'acq_expires']].max().reset_index()

    title_cols2exclude = [
        'project_type',
        'web_site',
        'copyright_holder',
        'producer',
        'aka_1',
        'original_language',
        'writer',
        'season',
        'number_of_episodes',
        #'link',
        'project_code',
        'logline',
        'rating',
        'short_synopsis',
        'original_format',
        'country_of_origin',
        'status',
        #'rating_usa',
        'number_of_seasons',
        'imdb'
    ]

    # merge free_tv_avails with the titles dataframe
    free_tv_avails = free_tv_avails.merge(
        titles.reset_index().loc[:,[col for col in titles.reset_index().columns if col not in title_cols2exclude]],
        how='inner',
        left_on='title',
        right_on='title'
    )

    # Free TV sales activity
    free_tv_sales = sales_activity.loc[pd.IndexSlice[:, 'Free TV', :, :], :].dropna(how='all', axis=1).reset_index()

    # drop the second level of the column index
    free_tv_sales.columns = free_tv_sales.columns.droplevel(1)

    # drop the 'right_name' column
    free_tv_sales.drop('right_name', axis=1, inplace=True)
    free_tv_sales.drop('group', axis=1, inplace=True)

    # group by Title and country_name, return the max value for each column exluding NaNs
    free_tv_sales = free_tv_sales.groupby(['title', 'country_name']).max()
    free_tv_sales.reset_index(inplace=True)

    free_tv_sales_cols = free_tv_sales.columns.tolist()
    free_tv_sales_cols.remove('title')
    free_tv_sales_cols.remove('country_name')

    # merge free_tv_avails with free_tv_sales
    free_tv_avails = free_tv_avails.merge(
        free_tv_sales,
        how='left',
        left_on=['title', 'country_name'],
        right_on=['title', 'country_name']
    )

    # create a 'First Run / Library' column in free_tv_avails with the value 'First Run' if all the values in the free_tv_sales_cols are NaN, otherwise 'Library'
    free_tv_avails['First Run / Library'] = free_tv_avails.loc[:, free_tv_sales_cols].isna().all(axis=1).map({True: 'First Run', False: 'Library'})

    free_tv_avails.drop('title', axis=1, inplace=True)
    free_tv_avails.rename(columns={'name': 'title'}, inplace=True)

    # reorder the columns of free_tv_avails. Put 'title', 'country_name', 'First Run / Library', 'exclusive', 'acq_expires' first and then the rest of the columns. 
    # Start by getting the index of the columns
    free_tv_avails_cols = free_tv_avails.columns.tolist()

    reorder_cols = [    
        'title', 
        'country_name', 
        'year_completed', 
        'First Run / Library', 
        'exclusive', 
        'acq_expires', 
        'genre',
        'cast',
        'director',
        'synopsis',
        'running_time',
        'dialogue_language',
        'subtitle_language',
        'website',
        'link',
        'password',
    ]

    for col in reorder_cols:
        free_tv_avails_cols.remove(col)
        
    # reorder the columns
    free_tv_avails = free_tv_avails[reorder_cols + free_tv_avails_cols]

    # sort free_tv_avails by 'exclusive' in descending order
    free_tv_avails.sort_values(['exclusive', 'country_name', 'title'], ascending=True, inplace=True)

    # rename the 'year_completed' column to 'year'
    free_tv_avails.rename(columns={
        'year_completed': 'year',
        'exclusive': 'start_date',
        }, inplace=True)

    # replace underscores with spaces and Title case the column names of free_tv_avails
    free_tv_avails.columns = [col.replace('_', ' ').title() for col in free_tv_avails.columns]

    # clean the strings of the columns that are objects
    for col in free_tv_avails.select_dtypes(include=['object']).columns:
        free_tv_avails[col] = free_tv_avails[col].apply(clean_str)

    # save free_tv_avails to an excel file, create one worksheet per country and name the worksheet with the country name
    with pd.ExcelWriter(os.path.join(app_dir, 'avails', 'free_tv_avails.xlsx')) as writer:
        for country in free_tv_avails['Country Name'].unique():
            sheet_df = free_tv_avails.loc[free_tv_avails['Country Name'] == country, :].dropna(how='all', axis=1)
            sheet_df.drop('Country Name', axis=1, inplace=True)
            sheet_df.to_excel(writer, sheet_name=country, index=False)

    # open the workbook object
    wb3 = openpyxl.load_workbook(os.path.join(app_dir, 'avails', 'free_tv_avails.xlsx'))

    # Format the date columns of each sheet of 'free_tv_avails.xlsx' to be in the format 'YYYY-MM-DD'
    for sheet in wb3.sheetnames:
        ws3 = wb3[sheet]
        ws3.auto_filter.ref = ws3.dimensions
        ws3.freeze_panes = ws3['C2']
        for col in ws3.iter_cols(min_col=1):
            for cell in col:
                ws3.column_dimensions[cell.column_letter].width = 30
                if cell.column_letter in ['D', 'E']:
                    cell.number_format = 'YYYY-MM-DD'
        for row in ws3.iter_rows(min_row=1):
            for cell in row:
                cell.font = Font(size=11)
                cell.alignment = Alignment(wrap_text=True)
                ws3.row_dimensions[cell.row].height = 50
        for row in ws3.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
            for cell in row:
                if cell.column_letter == 'A':
                    cell.alignment = Alignment(horizontal='left')
                    cell.font = Font(bold=True)


    # save the workbook object
    wb3.save(os.path.join(app_dir, 'avails', 'free_tv_avails.xlsx'))

    # close the workbook object
    wb3.close()